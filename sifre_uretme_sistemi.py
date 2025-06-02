
import random
import string
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os
import win32com.client  # Sadece Windows için çalışır durumda

EXCEL_DOSYASI = "sifreler.xlsx"

def sifre_uret(uzunluk, buyuk_harf=True, kucuk_harf=True, rakam=True, sembol=True):
    karakter_havuzu = ''
    if buyuk_harf:
        karakter_havuzu += string.ascii_uppercase
    if kucuk_harf:
        karakter_havuzu += string.ascii_lowercase
    if rakam:
        karakter_havuzu += string.digits
    if sembol:
        karakter_havuzu += string.punctuation
    if not karakter_havuzu:
        return "En az bir karakter türü seçmelisiniz!"
    return ''.join(random.choice(karakter_havuzu) for _ in range(uzunluk))

def excel_dosyasi_var_mi(dosya_adi):
    return os.path.exists(dosya_adi)# dosya var mı kontrol ediyor

def sifreyi_kaydet_excel(kullanim_adi, sifre, dosya_adi=EXCEL_DOSYASI):
    zaman = datetime.now().strftime("%Y-%m-%d %H:%M:%S")#sırayla tarih bilgilerini alıyor  yıl-ay-gün-saat-dakika-saniye
    if not excel_dosyasi_var_mi(dosya_adi):
        wb = Workbook()
        ws = wb.active 
        ws.append(["Kullanım", "Şifre", "Tarih/Saat"]) #aktif excel dosyasındaki sütün bilgilerini alıyor
    else:
        wb = load_workbook(dosya_adi)
        ws = wb.active

    ws.append([kullanim_adi, sifre, zaman])
    wb.save(dosya_adi)
    print(f" Şifre '{dosya_adi}' dosyasına kaydedildi.")

    os.startfile(dosya_adi)

def excel_dosyasina_parola_ekle(dosya_yolu, parola):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(os.path.abspath(dosya_yolu))
    wb.Password = parola
    wb.SaveAs(os.path.abspath(dosya_yolu), None, Password=parola)
    wb.Close()
    excel.Quit()
    print(f" Excel dosyasına parola uygulandı: {parola}")


print(" Güvenli Şifre Üretici (Excel'e Kayıt + Parola)\n")

try:
    uzunluk = int(input("Şifre uzunluğunu girin: "))
except ValueError:
    print("Lütfen geçerli bir sayı girin.")
    exit()

buyuk = input("Büyük harf olsun mu? (e/h): ").strip().lower() == 'e'
kucuk = input("Küçük harf olsun mu? (e/h): ").strip().lower() == 'e'
rakam = input("Rakam olsun mu? (e/h): ").strip().lower() == 'e'
sembol = input("Sembol olsun mu? (e/h): ").strip().lower() == 'e'

sifre = sifre_uret(uzunluk, buyuk, kucuk, rakam, sembol)
print(f"\n Oluşturulan Şifre: {sifre}")

kaydet = input("Bu şifreyi Excel dosyasına kaydetmek ister misin? (e/h): ").strip().lower() == 'e' #.strip gereksiz yer kaplayan elemanları siler
if kaydet:
    kullanim = input("Bu şifre nerede kullanılacak? (örn: Instagram, e-Devlet): ").strip()
    sifreyi_kaydet_excel(kullanim, sifre)
    #excel dosyasına girişte paralo sorsun istemezseniz bu satırları kullanıma kapatabilirsiniz

    parola_sor = input("Excel dosyasını parola ile korumak ister misin? (e/h): ").strip().lower() == 'e'
    if parola_sor:
        parola_icerigi = input("Parola ne olsun?: ").strip()
        excel_dosyasina_parola_ekle(EXCEL_DOSYASI, parola_icerigi)
else:
    print(" Şifre kaydedilmedi.")
